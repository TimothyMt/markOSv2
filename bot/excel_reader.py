"""
Excel read-back — parse user-edited .xlsx uploads back into the session.

Flow:
  1. Bot exports skill output as .xlsx (filename = "{skill}_{business}.xlsx").
  2. User edits the file (e.g. in Google Sheets / Excel) and uploads it back.
  3. detect_skill_from_filename() maps the filename → originating skill.
  4. read_xlsx_to_markdown() reconstructs sheet content as markdown tables.
  5. Handler stores it as a new version of that skill's result so downstream
     skills (content production, etc.) follow the user's edits.

Detection is FILENAME-based (cách 1): the user must keep the original filename
so the skill can be identified. We surface that reminder when sending Excel.
"""
import io
import logging
import re
from typing import Optional

logger = logging.getLogger(__name__)

# Max cells/rows we read to avoid runaway memory on huge uploads.
_MAX_ROWS = 400
_MAX_COLS = 30


def detect_skill_from_filename(filename: str) -> Optional[str]:
    """Map an uploaded filename back to the skill that produced it.

    Bot exports use stem "{task_name}" or "{task_name}_{business_slug}".
    We match the LONGEST known task name that the stem starts with, so
    'video_script_gen_Shop.xlsx' matches 'video_script_gen', not 'video'.
    Returns the task_name or None if no match.
    """
    if not filename:
        return None
    stem = re.sub(r"\.(xlsx|xlsm|xls)$", "", filename.strip(), flags=re.IGNORECASE)
    if not stem:
        return None

    from agents.task_registry import TASK_REGISTRY

    # Longest names first → avoid short-prefix false matches.
    names = sorted(TASK_REGISTRY.keys(), key=len, reverse=True)
    for name in names:
        if stem == name or stem.startswith(name + "_"):
            return name
    return None


def _cell_to_str(value) -> str:
    """Render a cell value as a clean string for markdown."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("|", "/").replace("\n", " ").strip()


def _sheet_to_markdown(ws) -> str:
    """Convert a worksheet to a markdown table (first non-empty row = header)."""
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx >= _MAX_ROWS:
            break
        cells = [_cell_to_str(c) for c in row[:_MAX_COLS]]
        # Drop trailing empties
        while cells and cells[-1] == "":
            cells.pop()
        if any(cells):
            rows.append(cells)

    if not rows:
        return ""

    # Normalize width to the widest row
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]

    header = rows[0]
    # Blank header cells → generic column names
    header = [h or f"Cột {i+1}" for i, h in enumerate(header)]

    lines = ["| " + " | ".join(header) + " |"]
    lines.append("|" + "|".join(["---"] * width) + "|")
    for r in rows[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def read_xlsx_to_markdown(xlsx_bytes: bytes) -> str:
    """Read an .xlsx file → markdown. Each sheet becomes a '## {sheet}' table.

    Returns markdown string (may be empty if file has no usable content).
    Raises on unreadable / non-xlsx bytes — caller should catch.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    parts = []
    try:
        for ws in wb.worksheets:
            md = _sheet_to_markdown(ws)
            if not md:
                continue
            title = (ws.title or "Sheet").strip()
            # Skip noisy default sheet names if only one sheet anyway
            if len(wb.worksheets) > 1:
                parts.append(f"## {title}\n\n{md}")
            else:
                parts.append(md)
    finally:
        wb.close()

    return "\n\n".join(parts).strip()
