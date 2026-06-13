"""
Output renderers — convert parsed agent output to deliverable formats.

Per-skill primary_deliverable determines which format is sent as main attachment:
  - HTML: all skills support (general purpose, mobile-safe)
  - EXCEL: content_calendar (table grid)
  - MARKDOWN: ad_copy, video_scripts, briefs (deliverable for downstream tools)

HTML always generated as fallback. Excel/Markdown generated when primary_deliverable matches.
"""
import io
import os
import re
import logging
from datetime import datetime
from typing import Optional

from agents.skills import OutputFormat

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────
# Template-based Excel output
# ─────────────────────────────────────────────────────────────────

_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "content_generation_template.xlsx",
)

# Maps skill_name → sheet name inside the template workbook.
# Skills listed here will fill the corresponding sheet instead of
# building a new workbook from scratch.
SKILL_TEMPLATE_SHEET: dict[str, str] = {
    "content_generator":   "📅 Content Calendar",
    # content_calendar KHÔNG map vào template cố định: nó là bảng KẾ HOẠCH
    # (Funnel / Nhóm khách / Topic / Hook angle...) — cột khác hẳn template
    # post 15-cột → ép vào template chỉ ra file rỗng. Dùng dynamic extraction
    # (render_excel_file) để dựng workbook từ chính các bảng kế hoạch của nó.
    "post_batch":          "📅 Content Calendar",   # Content Suite v2
    "ads_generator":       "✍️ Ad Copy",
    "ads_copy":            "✍️ Ad Copy",
    # video_scripts / video_script_gen KHÔNG map vào template "🎬 Video Script":
    # prompt xuất bảng 8 cột linh hoạt (Version / Creator Type / Platform / Framework /
    # Beat Breakdown / Visual Direction / Caption Hook + Hashtags / Ghi chú) trong khi
    # sheet template lại tách 5-act cố định (Hook 3s / Problem / Solution / Social Proof /
    # CTA / Music/SFX) → match theo tên cột chỉ trùng 4 cột, toàn bộ lời thoại (Beat
    # Breakdown) bị rớt → file rỗng. Dùng dynamic extraction (render_excel_file) để dựng
    # workbook đúng theo bảng LLM xuất ra.
    "ugc_brief":           "🤝 UGC Brief",
    "email_zalo_sequence": "📧 Email & Zalo",
}


# ─────────────────────────────────────────────────────────────────
# Parsers — handle 3 output format variants
# ─────────────────────────────────────────────────────────────────

def parse_strategic_output(text: str) -> dict:
    """Parse STRATEGIC_4_SECTION output into {insight, summary, benchmarks, detail}."""
    from bot.html_report import parse_agent_output
    return parse_agent_output(text)


def parse_operational_deliverable(text: str) -> dict:
    """Parse OPERATIONAL_DELIVERABLE output into {summary, deliverable, raw}."""
    # ALWAYS preserve raw để Excel renderer có thể fallback
    result = {"summary": "", "deliverable": "", "raw": text}

    # Match "## 🎯 Tóm tắt nhanh" section
    summary_match = re.search(
        r"##\s*🎯[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        text, flags=re.DOTALL
    )
    if summary_match:
        result["summary"] = summary_match.group(1).strip()

    # Match "## 📄 Deliverable" section
    deliverable_match = re.search(
        r"##\s*📄[^\n]*\n+(.*?)(?=\Z)",
        text, flags=re.DOTALL
    )
    if deliverable_match:
        result["deliverable"] = deliverable_match.group(1).strip()

    # Fallback: nếu parse fail, dùng cả text làm deliverable
    if not result["summary"] and not result["deliverable"]:
        result["deliverable"] = text.strip()

    return result


def parse_operational_analysis(text: str) -> dict:
    """Parse OPERATIONAL_ANALYSIS output into {summary, kpi_table, root_cause, actions, forecast}."""
    result = {"summary": "", "kpi_table": "", "root_cause": "", "actions": "", "forecast": "", "raw": text}

    patterns = {
        "summary":    r"##\s*📊\s*Tóm tắt[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "kpi_table":  r"##\s*📈[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "root_cause": r"##\s*🔬[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "actions":    r"##\s*🎯[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "forecast":   r"##\s*📉[^\n]*\n+(.*?)(?=\Z)",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text, flags=re.DOTALL)
        if m:
            result[key] = m.group(1).strip()

    if not any(v for k, v in result.items() if k != "raw"):
        result["summary"] = text.strip()[:2000]

    return result


def parse_by_format(text: str, output_format: OutputFormat) -> dict:
    """Dispatch parser based on output format."""
    if output_format == OutputFormat.STRATEGIC_4_SECTION:
        return parse_strategic_output(text)
    elif output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        return parse_operational_deliverable(text)
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        return parse_operational_analysis(text)
    return {"raw": text}


# ─────────────────────────────────────────────────────────────────
# Telegram card formatters
# ─────────────────────────────────────────────────────────────────

def format_telegram_card(
    skill_name: str,
    skill_label: str,
    skill_emoji: str,
    parsed: dict,
    output_format: OutputFormat,
    file_attached_hint: Optional[str] = None,
) -> str:
    """Build Telegram preview card. Long content always goes to HTML/file."""
    header = f"*{skill_emoji} {skill_label.upper()}*"
    separator = "━" * 25
    parts = [header, separator, ""]

    if output_format == OutputFormat.STRATEGIC_4_SECTION:
        if parsed.get("insight"):
            insight = parsed["insight"].strip().strip('"').strip("'")
            parts.append("💡 *Insight quan trọng nhất:*")
            parts.append(f"_{insight}_")
            parts.append("")
        if parsed.get("summary"):
            parts.append("📌 *Tóm tắt:*")
            parts.append(parsed["summary"].strip())
            parts.append("")
        if parsed.get("benchmarks"):
            parts.append("📊 *Benchmarks:*")
            parts.append(parsed["benchmarks"].strip())
            parts.append("")

    elif output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        if parsed.get("summary"):
            parts.append("🎯 *Tóm tắt nhanh:*")
            parts.append(parsed["summary"].strip())
            parts.append("")

    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        if parsed.get("summary"):
            parts.append("📊 *Tổng quan:*")
            parts.append(parsed["summary"].strip())
            parts.append("")

    # Hint about attached file
    if file_attached_hint:
        parts.append(f"📎 _{file_attached_hint}_")
    else:
        parts.append("📎 _Xem chi tiết trong file đính kèm bên dưới_")

    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────────
# Markdown deliverable file generator
# ─────────────────────────────────────────────────────────────────

def render_markdown_file(
    skill_name: str,
    skill_label: str,
    parsed: dict,
    output_format: OutputFormat,
    business_name: str = "",
) -> bytes:
    """Render skill output as a .md file for download (designer/dev/creator workflow)."""
    lines = [
        f"# {skill_label} — {business_name or 'Marketing OS'}",
        f"*Generated by Max — AI CMO · {datetime.now().strftime('%d/%m/%Y · %H:%M')}*",
        "",
        "---",
        "",
    ]

    if output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        if parsed.get("summary"):
            lines += ["## 🎯 Tóm tắt nhanh", "", parsed["summary"].strip(), "", "---", ""]
        if parsed.get("deliverable"):
            lines += [parsed["deliverable"].strip()]
    elif output_format == OutputFormat.STRATEGIC_4_SECTION:
        if parsed.get("insight"):
            lines += [f"> 💡 **Insight:** {parsed['insight'].strip().strip(chr(34)).strip(chr(39))}", ""]
        if parsed.get("summary"):
            lines += ["## 🎯 Tóm tắt", "", parsed["summary"].strip(), ""]
        if parsed.get("benchmarks"):
            lines += ["## 📊 Benchmarks", "", parsed["benchmarks"].strip(), ""]
        if parsed.get("detail"):
            lines += ["## 📄 Phân tích chi tiết", "", parsed["detail"].strip(), ""]
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        # Dump all sections
        for key in ["summary", "kpi_table", "root_cause", "actions", "forecast"]:
            if parsed.get(key):
                lines += [f"## {key.replace('_', ' ').title()}", "", parsed[key].strip(), "", "---", ""]
    else:
        lines += [parsed.get("raw", "")]

    return "\n".join(lines).encode("utf-8")


# ─────────────────────────────────────────────────────────────────
# Excel renderer — for content_calendar
# ─────────────────────────────────────────────────────────────────

def _clean_cell(value) -> str:
    """Strip markdown chars (** __ *) khỏi cell content. Trả về str."""
    if value is None:
        return ""
    s = str(value)
    # Remove bold/italic markers
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
    s = re.sub(r"__(.+?)__", r"\1", s)
    s = re.sub(r"(?<!\*)\*(?!\*)([^*\n]+?)\*(?!\*)", r"\1", s)
    s = re.sub(r"(?<!_)_(?!_)([^_\n]+?)_(?!_)", r"\1", s)
    return s.strip()


def _safe_sheet_name(raw: str, idx: int, used: set) -> str:
    """Sheet name an toàn (≤31 chars, không có : \\ / ? * [ ], unique)."""
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", raw or "")
    # Remove emoji prefix to save chars
    cleaned = re.sub(r"^[^\w\d]+", "", cleaned).strip()
    cleaned = cleaned[:28].strip() or f"Sheet {idx+1}"
    # Ensure unique
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f" {n}"
        cleaned = base[:28 - len(suffix)] + suffix
        n += 1
    used.add(cleaned)
    return cleaned


def _is_keyvalue_table(headers: list, rows: list[list]) -> bool:
    """Detect mini key-value table: 2 columns, headers look generic, repeating field names."""
    if len(headers) != 2:
        return False
    GENERIC = {"field", "value", "key", "thông tin", "chi tiết", "metadata"}
    h0 = (headers[0] or "").lower().strip("* ")
    h1 = (headers[1] or "").lower().strip("* ")
    if h0 in GENERIC or h1 in GENERIC:
        return True
    # If first column values look like field names (Ngày, Kênh, Pillar...)
    FIELD_HINTS = {"ngày", "kênh", "pillar", "funnel", "source", "format", "hook",
                   "cta", "angle", "topic", "tier", "platform", "campaign"}
    first_col_vals = [_clean_cell(r[0]).lower() for r in rows[:5]]
    if sum(1 for v in first_col_vals if any(h in v for h in FIELD_HINTS)) >= 2:
        return True
    return False


def _pivot_keyvalue_tables(tables: list[tuple]) -> Optional[tuple]:
    """Convert nhiều mini key-value tables thành 1 master sheet.
    Mỗi mini-table thành 1 row, key = column header.
    Returns (title, headers, rows) hoặc None nếu không có cụm KV nào.
    """
    kv_tables = [(t, h, r) for (t, h, r) in tables if _is_keyvalue_table(h, r)]
    if len(kv_tables) < 2:
        return None

    # Collect all unique field names across all KV tables (preserve order from first)
    all_fields: list[str] = []
    seen_fields = set()
    for _, _, rows in kv_tables:
        for row in rows:
            if not row: continue
            field = _clean_cell(row[0])
            field_lower = field.lower()
            if field and field_lower not in seen_fields:
                seen_fields.add(field_lower)
                all_fields.append(field)

    # Build master rows: 1 row per table
    master_headers = ["Bài"] + all_fields
    master_rows = []
    for tbl_title, _, rows in kv_tables:
        # Strip emoji + clean title
        clean_title = _clean_cell(tbl_title or "")
        clean_title = re.sub(r"^[^\w\d]+", "", clean_title).strip()[:60]
        row_dict = {_clean_cell(r[0]).lower(): _clean_cell(r[1] if len(r) > 1 else "") for r in rows if r}
        master_rows.append([clean_title] + [row_dict.get(f.lower(), "") for f in all_fields])

    return ("📊 Tổng hợp", master_headers, master_rows)


def _split_table_by_week(headers: list, rows: list[list]) -> Optional[dict]:
    """Nếu table có cột 'Tuần' → group rows theo tuần.
    Returns {week_label: rows} hoặc None nếu không có cột Tuần.
    Dùng cho content_generator output để tạo sheet riêng cho từng tuần.
    """
    if not headers or not rows:
        return None
    # Tìm index cột 'Tuần' (case-insensitive)
    week_idx = None
    for i, h in enumerate(headers):
        h_clean = (h or "").strip().lower().lstrip("*").rstrip("*").strip()
        if h_clean in ("tuần", "tuan", "week", "tuần ", "wk"):
            week_idx = i
            break
    if week_idx is None:
        return None

    groups: dict[str, list] = {}
    for row in rows:
        if len(row) <= week_idx:
            continue
        week_label = _clean_cell(row[week_idx]).strip() or "Khác"
        # Normalize "Tuần 1" / "Tuan 1" / "Week 1"
        week_label = re.sub(r"(?i)^(tu[aâà]n|week|wk)\s*", "Tuần ", week_label).strip()
        groups.setdefault(week_label, []).append(row)

    if len(groups) < 2:
        return None  # Chỉ có 1 tuần, không cần split
    return groups


def _norm_header(s: str) -> str:
    """Normalize header for fuzzy column matching (strip punctuation, lowercase)."""
    return re.sub(r'[^a-z0-9À-ɏ]', '', s.lower()) if s else ""


def render_template_excel(
    skill_name: str,
    skill_label: str,
    parsed: dict,
    output_format: OutputFormat,
    business_name: str = "",
) -> Optional[bytes]:
    """Fill content_generation_template.xlsx with LLM data for the skill's mapped sheet.

    Falls back to render_excel_file() if: template file missing, skill not mapped,
    or no tables found in LLM output.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return render_excel_file(skill_name, skill_label, parsed, output_format, business_name)

    target_sheet = SKILL_TEMPLATE_SHEET.get(skill_name)
    if not target_sheet or not os.path.exists(_TEMPLATE_PATH):
        return render_excel_file(skill_name, skill_label, parsed, output_format, business_name)

    try:
        wb = load_workbook(_TEMPLATE_PATH)
    except Exception as exc:
        logger.warning("render_template_excel: failed to load template: %s", exc)
        return render_excel_file(skill_name, skill_label, parsed, output_format, business_name)

    if target_sheet not in wb.sheetnames:
        logger.warning("render_template_excel: sheet '%s' not found in template", target_sheet)
        return render_excel_file(skill_name, skill_label, parsed, output_format, business_name)

    ws = wb[target_sheet]

    # Build full_text from parsed — same logic as render_excel_file
    if output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        full_text = "\n\n".join(filter(None, [
            parsed.get("deliverable", ""),
            parsed.get("summary", ""),
            parsed.get("raw", ""),
        ]))
    else:
        full_text = "\n\n".join(
            parsed.get(k, "") for k in ["summary", "kpi_table", "root_cause", "actions", "forecast"]
        )
        if not full_text.strip():
            full_text = parsed.get("raw", "")

    tables = _extract_markdown_tables(full_text)
    if not tables:
        try:
            rebuilt = _haiku_rebuild_table(full_text, skill_name)
            if rebuilt:
                tables = _extract_markdown_tables(rebuilt)
        except Exception as exc:
            logger.warning("render_template_excel: Haiku rebuild failed: %s", exc)

    if not tables:
        logger.error("render_template_excel [%s]: no tables extracted — skipping", skill_name)
        return None

    # Template header row is row 3 (rows 1-2 are title + subtitle)
    tmpl_headers = [ws.cell(row=3, column=c).value for c in range(1, 16)]
    tmpl_norm_to_col = {_norm_header(h): i + 1 for i, h in enumerate(tmpl_headers) if h}

    # Pick table with highest column-overlap against template headers
    best_table = tables[0]
    best_score = 0
    for t in tables:
        _, hdrs, _ = t
        score = sum(1 for h in hdrs if _norm_header(h) in tmpl_norm_to_col)
        if score > best_score:
            best_score = score
            best_table = t

    _, headers, rows = best_table
    llm_norm_to_idx = {_norm_header(h): i for i, h in enumerate(headers) if h}

    # Pairs: (template column 1-indexed, llm column 0-indexed)
    col_pairs = [
        (tmpl_col, llm_norm_to_idx[norm])
        for norm, tmpl_col in tmpl_norm_to_col.items()
        if norm in llm_norm_to_idx
    ]

    body_font = Font(name="Arial", size=10)
    body_align = Alignment(vertical="top", wrap_text=True)

    # Clear existing data rows (keep rows 1-3: title, subtitle, headers)
    data_start = 4
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, 16):
            ws.cell(row=r, column=c).value = None

    # Fill data
    for r_idx, row_data in enumerate(rows):
        excel_row = data_start + r_idx
        for tmpl_col, llm_col in col_pairs:
            val = row_data[llm_col] if llm_col < len(row_data) else None
            cell = ws.cell(row=excel_row, column=tmpl_col)
            cell.value = _clean_cell(str(val)) if val is not None else None
            cell.font = body_font
            cell.alignment = body_align

    # Stamp business name into title if not already there
    if business_name:
        title_cell = ws["A1"]
        current = title_cell.value or ""
        if business_name not in current:
            title_cell.value = f"{current} — {business_name}"

    # Prune: send only the relevant sheet + instructions; drop the other 6 sheets
    KEEP = {target_sheet, "📖 Hướng dẫn"}
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in KEEP:
            del wb[sheet_name]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_excel_file(
    skill_name: str,
    skill_label: str,
    parsed: dict,
    output_format: OutputFormat,
    business_name: str = "",
    post_ids: Optional[list[str]] = None,
) -> Optional[bytes]:
    """Render skill output as .xlsx.
    Skills in SKILL_TEMPLATE_SHEET are routed to render_template_excel() which
    fills content_generation_template.xlsx instead of building a new workbook.
    Other skills use dynamic table extraction + openpyxl.

    post_ids: nếu truyền vào (content_calendar) — danh sách POST-XXX IDs theo
    đúng thứ tự parse_calendar_to_posts() đã gán, dùng để chèn cột "ID" đầu
    mỗi bảng dữ liệu (không áp cho overview/key-value sheet).
    """
    if skill_name in SKILL_TEMPLATE_SHEET:
        return render_template_excel(skill_name, skill_label, parsed, output_format, business_name)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — falling back to no Excel export")
        return None

    _PILLAR_FILLS = {
        "educate":  PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "trust":    PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "engage":   PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),
        "convert":  PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    }

    def _pillar_fill(value: str):
        return _PILLAR_FILLS.get(value.strip().lower()) if value else None

    if output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        # Include raw để cover case LLM output table ngoài section "Deliverable"
        full_text = "\n\n".join(filter(None, [
            parsed.get("deliverable", ""),
            parsed.get("summary", ""),
            parsed.get("raw", ""),
        ]))
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        full_text = "\n\n".join(
            parsed.get(k, "") for k in ["summary", "kpi_table", "root_cause", "actions", "forecast"]
        )
        # Fallback to raw nếu structured parse fail
        if not full_text.strip():
            full_text = parsed.get("raw", "")
    else:
        full_text = parsed.get("detail", "") + "\n\n" + parsed.get("raw", "")

    tables = _extract_markdown_tables(full_text)
    if not tables:
        logger.warning("render_excel_file [%s]: no markdown tables found via parser, trying Haiku rebuild (full_text len=%d)",
                       skill_name, len(full_text))
        # Last resort: dùng Haiku rebuild table format
        try:
            rebuilt = _haiku_rebuild_table(full_text, skill_name)
            if rebuilt:
                tables = _extract_markdown_tables(rebuilt)
                if tables:
                    logger.info("render_excel_file [%s]: Haiku rebuild succeeded → %d tables", skill_name, len(tables))
        except Exception as e:
            logger.warning("Haiku rebuild failed: %s", e)

        if not tables:
            logger.error("render_excel_file [%s]: still no tables after Haiku rebuild. First 500 chars: %s",
                          skill_name, full_text[:500])
            return None

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial")
    body_align = Alignment(vertical="top", wrap_text=True)
    title_font = Font(bold=True, size=14, name="Arial")

    # Detect & merge key-value tables
    overview = _pivot_keyvalue_tables(tables)
    if overview:
        sheets_to_render = [overview]
        # Add non-KV tables
        for t in tables:
            if not _is_keyvalue_table(t[1], t[2]):
                sheets_to_render.append(t)
    else:
        sheets_to_render = tables[:8]  # cap 8 sheets

    # Content Calendar — chèn cột "ID" (POST-XXX) đầu mỗi bảng dữ liệu, khớp
    # thứ tự với parse_calendar_to_posts() để /post <ID> trỏ đúng row trong Excel.
    if skill_name == "content_calendar" and post_ids:
        new_sheets = []
        post_idx = 0
        for t_title, t_headers, t_rows in sheets_to_render:
            if _is_keyvalue_table(t_headers, t_rows):
                new_sheets.append((t_title, t_headers, t_rows))
                continue
            new_headers = ["ID"] + list(t_headers)
            new_rows = []
            for row in t_rows:
                pid = post_ids[post_idx] if post_idx < len(post_ids) else ""
                post_idx += 1
                new_rows.append([pid] + list(row))
            new_sheets.append((t_title, new_headers, new_rows))
        sheets_to_render = new_sheets

    # SPECIAL — Content Generator: chỉ giữ MASTER table (có cột Tuần + Bài),
    # bỏ qua các mini-table phụ trong content (size guides, comparison tables, etc.)
    if skill_name == "content_generator":
        master_table = None
        for t_title, t_headers, t_rows in tables:
            cleaned_headers_lower = [_clean_cell(h).lower().strip() for h in t_headers]
            has_tuan = any(h in ("tuần", "tuan", "week") for h in cleaned_headers_lower)
            has_bai = any(h in ("bài", "bai", "post", "#") for h in cleaned_headers_lower)
            if has_tuan and has_bai:
                master_table = (t_title, t_headers, t_rows)
                break

        if master_table:
            s_title, s_headers, s_rows = master_table
            sheets_to_render = []
            # Overview sheet
            sheets_to_render.append((f"📊 Tổng hợp ({len(s_rows)} bài)", s_headers, s_rows))
            # Split by week
            week_groups = _split_table_by_week(s_headers, s_rows)
            if week_groups:
                for week_label in sorted(week_groups.keys(), key=lambda x: (len(x), x)):
                    week_rows = week_groups[week_label]
                    sheets_to_render.append((f"{week_label} ({len(week_rows)} bài)", s_headers, week_rows))
        else:
            # Không có master table → LLM output thiếu. Vẫn render những gì có để debug
            logger.warning("content_generator: master table (Tuần+Bài columns) not found in output. Falling back to default render.")

    used_names = set()
    for idx, (table_title, headers, rows) in enumerate(sheets_to_render):
        raw_name = table_title or f"Bảng {idx+1}"
        sheet_name = _safe_sheet_name(raw_name, idx, used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        clean_title = _clean_cell(table_title or "")
        if clean_title:
            ws.append([clean_title])
            ws["A1"].font = title_font
            ws.append([])

        # Headers
        clean_headers = [_clean_cell(h) for h in headers]
        header_row = ws.max_row + 1 if ws.max_row > 0 else 1
        ws.append(clean_headers)
        for col_idx in range(1, len(clean_headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Detect pillar column index (0-based) for row coloring
        pillar_col_idx = next(
            (i for i, h in enumerate(clean_headers)
             if h.strip().lower() in ("pillar", "content pillar")),
            None,
        )

        # Data rows — strip markdown
        for row in rows:
            cleaned_row = [_clean_cell(c) for c in row]
            ws.append(cleaned_row)
            r_idx = ws.max_row
            row_fill = (
                _pillar_fill(cleaned_row[pillar_col_idx])
                if pillar_col_idx is not None and pillar_col_idx < len(cleaned_row)
                else None
            )
            for c_idx in range(1, len(cleaned_row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = body_font
                cell.alignment = body_align
                if row_fill:
                    cell.fill = row_fill

        # Auto column width (capped 60)
        for col_idx in range(1, len(clean_headers) + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(
                len(clean_headers[col_idx-1]) if col_idx-1 < len(clean_headers) else 0,
                *[len(str(r[col_idx-1] or "")) for r in rows if col_idx-1 < len(r)],
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _haiku_rebuild_table(text: str, skill_name: str) -> Optional[str]:
    """Khi parser không tìm thấy pipe table, dùng Haiku để convert
    output (dạng bullet/list/narrative) thành markdown table chuẩn.
    Returns: chuỗi chứa markdown table hoặc None.
    """
    if not text or len(text) < 100:
        return None

    try:
        import anthropic
        from config import CLAUDE_HAIKU_MODEL, ANTHROPIC_API_KEY
    except ImportError:
        return None

    # Skill-specific schema
    schemas = {
        "content_generator": (
            "Bài content output với 15 cột: "
            "Tuần | Bài | Ngày | Kênh | Pillar | Funnel | Source | Format | "
            "Angle | Hook | Body | CTA | Hashtags | Visual | Status"
        ),
        "content_calendar": (
            "Lịch nội dung với cột: "
            "Tuần | Ngày | Kênh | Pillar | Funnel | Nhóm khách | Source | Hook angle | Topic | Format | Owner"
        ),
    }
    schema_desc = schemas.get(skill_name, "Tự deduce cấu trúc từ output")

    system_prompt = f"""Convert output dưới đây thành 1 markdown pipe table.

Schema mong đợi: {schema_desc}

QUY TẮC:
- Output CHỈ markdown table, không có text khác
- Format: `| col1 | col2 | ... |`
- Có separator: `|---|---|...|`
- KHÔNG dùng dấu | trong cell (thay bằng /)
- Nếu output gốc có nhiều bài/items → mỗi bài 1 row
- Cell trống → để "..." không để rỗng"""

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model=CLAUDE_HAIKU_MODEL,
            max_tokens=8000,
            system=system_prompt,
            messages=[{"role": "user", "content": text[:30000]}],
        )
        return response.content[0].text.strip()
    except Exception as e:
        logger.warning("Haiku rebuild table failed: %s", e)
        return None


def _extract_markdown_tables(text: str) -> list[tuple[str, list[str], list[list[str]]]]:
    """Extract markdown tables from text. Returns list of (title, headers, rows).
    Title is the nearest preceding heading (###/####).
    Lenient: handles unicode pipes (｜), code blocks, indented tables."""
    if not text:
        return []

    # Normalize unicode pipes + remove code fences
    text = text.replace("｜", "|").replace("∣", "|")
    text = re.sub(r"```[a-zA-Z]*\n?", "", text)  # strip ```python etc.
    text = text.replace("```", "")

    tables = []
    lines = text.split("\n")

    current_title = ""
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Track title (last seen heading)
        h_match = re.match(r"^#{1,5}\s+(.+)$", line)
        if h_match:
            current_title = h_match.group(1).strip()

        # Detect table header row: starts with | and has at least 1 more |
        # More lenient: allow leading/trailing whitespace, missing trailing |
        if line.startswith("|") and line.count("|") >= 2 and i + 1 < len(lines):
            sep_line = lines[i + 1].strip()
            # Separator row: lenient — chỉ cần có | và --- pattern
            if sep_line.startswith("|") and "---" in sep_line and re.match(r"^\|[\s:|\-]+\|?$", sep_line):
                # Ensure trailing | (in case missing)
                header_line = line if line.endswith("|") else line + "|"
                headers = [c.strip() for c in header_line.strip("|").split("|")]
                rows = []
                j = i + 2
                while j < len(lines):
                    row_line = lines[j].strip()
                    # Lenient row match: starts with | and has at least 1 more |
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    # Ensure trailing |
                    if not row_line.endswith("|"):
                        row_line += "|"
                    row_cells = [c.strip() for c in row_line.strip("|").split("|")]
                    # Pad to match header length
                    while len(row_cells) < len(headers):
                        row_cells.append("")
                    rows.append(row_cells[:len(headers)])
                    j += 1
                if rows:
                    tables.append((current_title, headers, rows))
                i = j
                continue
        i += 1

    return tables
